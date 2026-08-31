import json
import sqlite3
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from openpyxl import Workbook

DB_PATH = Path(r"D:\Bible project\mapping\bible_map.db")
PORT = 3456
VIEWER_DIR = Path(r"D:\Bible project\mapping\web_templates")
STATIC_DIR = Path(r"D:\Bible project\mapping\web_static")
VIEWER_HTML = (
    (VIEWER_DIR / "index.html").read_text(encoding="utf-8")
    if VIEWER_DIR.exists()
    else "<h1>Bible Mapping Viewer - missing template</h1>"
)
CANONICAL_BOOKS = [
    "Genesis", "Exodus", "Leviticus", "Numbers", "Deuteronomy",
    "Joshua", "Judges", "Ruth", "1 Samuel", "2 Samuel",
    "1 Kings", "2 Kings", "1 Chronicles", "2 Chronicles",
    "Ezra", "Nehemiah", "Esther", "Job", "Psalms", "Proverbs",
    "Ecclesiastes", "Song of Solomon", "Isaiah", "Jeremiah",
    "Lamentations", "Ezekiel", "Daniel", "Hosea", "Joel",
    "Amos", "Obadiah", "Jonah", "Micah", "Nahum", "Habakkuk",
    "Zephaniah", "Haggai", "Zechariah", "Malachi",
    "Matthew", "Mark", "Luke", "John", "Acts", "Romans",
    "1 Corinthians", "2 Corinthians", "Galatians", "Ephesians",
    "Philippians", "Colossians", "1 Thessalonians", "2 Thessalonians",
    "1 Timothy", "2 Timothy", "Titus", "Philemon", "Hebrews",
    "James", "1 Peter", "2 Peter", "1 John", "2 John", "3 John",
    "Jude", "Revelation",
]

CONTENT_TYPES = {
    ".html": "text/html; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
}


class BibleAPIHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        con = None
        try:
            raw = self.path
            query = ""
            if "?" in raw:
                raw, query = raw.split("?", 1)
            path = urllib.parse.unquote(raw)
            params = urllib.parse.parse_qs(query)

            if path == "/" or path.startswith("/index.html"):
                html = VIEWER_HTML.encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(html)))
                self.end_headers()
                self.wfile.write(html)
                return

            api_path = path
            if api_path.startswith("/api/"):
                api_path = api_path[len("/api"):]
            elif api_path == "/api":
                api_path = ""

            for folder in (STATIC_DIR, Path("D:/Bible project/mapping/web_static")):
                candidate = folder / path.lstrip("/")
                if candidate.exists() and candidate.is_file():
                    content = candidate.read_bytes()
                    ctype = CONTENT_TYPES.get(
                        candidate.suffix.lower(), "application/octet-stream"
                    )
                    self.send_response(200)
                    self.send_header("Content-Type", ctype)
                    self.send_header("Content-Length", str(len(content)))
                    self.end_headers()
                    self.wfile.write(content)
                    return

            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()

            if api_path in ("/health", "/healthz", "/.well-known/health"):
                self.send_json(
                    200,
                    {"status": "ok", "service": "Bible Mapping API", "db": str(DB_PATH)},
                    con,
                )
                return

            if api_path == "/stats":
                stats = {
                    "total_verses": cur.execute("SELECT COUNT(*) FROM verses").fetchone()[0],
                    "total_words": cur.execute(
                        "SELECT COUNT(*) FROM word_occurrences"
                    ).fetchone()[0],
                    "total_books": cur.execute(
                        "SELECT COUNT(DISTINCT book) FROM verses"
                    ).fetchone()[0],
                    "lexicon_entries": cur.execute(
                        "SELECT COUNT(*) FROM lexicon_entries"
                    ).fetchone()[0],
                }
                self.send_json(200, stats, con)
                return

            if api_path == "/books":
                rows = cur.execute(
                    """
                    SELECT book,
                           COUNT(*) as verse_count,
                           COUNT(DISTINCT chapter) as chapters,
                           MIN(verse_id) as sample_verse_id
                    FROM verses
                    GROUP BY book
                    ORDER BY book
                    """
                ).fetchall()
                book_order = {name: idx for idx, name in enumerate(CANONICAL_BOOKS, start=1)}
                result = []
                for r in rows:
                    book, verse_count, chapters, sample_verse_id = r
                    book_code = (
                        sample_verse_id.split("##")[0]
                        if sample_verse_id
                        else book.upper().replace(" ", "")
                    )
                    result.append(
                        {
                            "book": book,
                            "verse_count": verse_count,
                            "chapters": chapters,
                            "book_code": book_code,
                            "book_order": book_order.get(book, 9999),
                        }
                    )
                result.sort(key=lambda r: (r["book_order"], r["book"]))
                self.send_json(200, result, con)
                return

            if api_path.startswith("/verse") or api_path == "/verse":
                verse_id = (
                    params.get("verse_id", [""])[0] if "verse_id" in params else ""
                )
                path_parts = api_path.split("/")
                path_has_full_verse = (
                    api_path.startswith("/verse/")
                    and len(path_parts) == 3
                    and "##" in path_parts[2]
                )
                if not verse_id and api_path.startswith("/verse/"):
                    candidate = urllib.parse.unquote(
                        api_path[len("/verse/"):]
                    )
                    if "##" in candidate:
                        verse_id = candidate
                        path_has_full_verse = True
                if not verse_id:
                    self.send_json(400, {"error": "Missing verse_id"}, con)
                    return
                parts = [c for c in verse_id.split("#") if c]
                if len(parts) >= 3:
                    verse_id = "##".join(parts[:3])
                is_full_verse_lookup = path_has_full_verse or (
                    api_path == "/verse"
                    and len(parts) == 3
                    and parts[1].isdigit()
                    and parts[2].isdigit()
                )
                if (
                    not is_full_verse_lookup
                    and len(parts) >= 2
                    and parts[1].isdigit()
                ):
                    book_code = parts[0]
                    chapter = int(parts[1])
                    query = "SELECT * FROM verses WHERE verse_id LIKE ? AND chapter=?"
                    rows = cur.execute(
                        query, (f"{book_code}##%", chapter)
                    ).fetchall()
                    self.send_json(
                        200,
                        {"rows": [dict(r) for r in rows], "verse_count": len(rows)},
                        con,
                    )
                    return
                rows = cur.execute(
                    "SELECT * FROM verses WHERE verse_id=?", (verse_id,)
                ).fetchall()
                translations = {}
                for r in rows:
                    key = r["translation_id"]
                    translations[key] = dict(r)
                self.send_json(200, {"translations": translations}, con)
                return

            if api_path == "/export":
                self._export_excel(con, cur)
                return

            if api_path.startswith("/compare/") or api_path == "/compare":
                verse_id = (
                    params.get("verse_id", [""])[0] if "verse_id" in params else ""
                )
                if not verse_id and api_path.startswith("/compare/"):
                    verse_id = urllib.parse.unquote(
                        api_path[len("/compare/"):]
                    )
                if not verse_id:
                    self.send_json(400, {"error": "Missing verse_id"}, con)
                    return
                parts = [c for c in verse_id.split("#") if c]
                if len(parts) >= 3:
                    verse_id = "##".join(parts[:3])
                rows = cur.execute(
                    "SELECT * FROM verses WHERE verse_id=?", (verse_id,)
                ).fetchall()
                mappings = cur.execute(
                    """
                    SELECT m.*, l.lemma, l.definition
                    FROM alignment_edges m
                    LEFT JOIN lexicon_entries l ON m.strongs_id = l.id
                    WHERE m.verse_id=? AND m.strongs_id IS NOT NULL
                    ORDER BY m.id
                    """,
                    (verse_id,),
                ).fetchall()
                result = {
                    "verse_id": verse_id,
                    "translations": {r["translation_id"]: dict(r) for r in rows},
                    "strongs_links": [dict(m) for m in mappings],
                }
                self.send_json(200, result, con)
                return

            if api_path.startswith("/strong/"):
                strongs_id = api_path.split("/")[2].upper()
                entry = cur.execute(
                    "SELECT * FROM lexicon_entries WHERE id=?", (strongs_id,)
                ).fetchone()
                words = cur.execute(
                    """
                    SELECT w.word_id, w.verse_id, w.translation_id, w.surface, w.position,
                           v.book, v.chapter, v.verse, v.text
                    FROM word_occurrences w
                    JOIN verses v ON w.verse_id = v.verse_id AND w.translation_id = v.translation_id
                    WHERE w.strongs_id=?
                    ORDER BY v.verse_id, w.position
                    LIMIT 100
                    """,
                    (strongs_id,),
                ).fetchall()
                occurrences = []
                for w in words:
                    occurrences.append({
                        "word_id": w[0],
                        "verse_id": w[1],
                        "translation_id": w[2],
                        "surface": w[3],
                        "position": w[4],
                        "reference": f"{w[5]} {w[6]}:{w[7]}",
                        "text": w[8],
                    })
                result = {
                    "id": strongs_id,
                    "lemma": entry["lemma"] if entry else None,
                    "definition": entry["definition"] if entry else None,
                    "root_form": entry["root_form"] if entry else None,
                    "language": entry["language"] if entry else None,
                    "occurrences": occurrences,
                }
                self.send_json(200, result, con)
                return

            if api_path.startswith("/trilingual"):
                verse_id = params.get("verse_id", [""])[0].strip()
                if not verse_id:
                    self.send_json(400, {"error": "Missing verse_id"}, con)
                    return
                parts = [c for c in verse_id.split("#") if c]
                if len(parts) >= 3:
                    verse_id = "##".join(parts[:3])
                rows = cur.execute(
                    "SELECT word_id, surface, strongs_id, translation_id FROM word_occurrences WHERE verse_id=? ORDER BY translation_id, position",
                    (verse_id,),
                ).fetchall()
                strongs_ids = sorted({r[2] for r in rows if r[2]})
                strongs_map = {}
                if strongs_ids:
                    placeholders = ",".join("?" for _ in strongs_ids)
                    for row in cur.execute(
                        f"SELECT id, language, lemma, definition FROM lexicon_entries WHERE id IN ({placeholders})",
                        strongs_ids,
                    ).fetchall():
                        strongs_map[row[0]] = {"language": row[1], "lemma": row[2], "definition": row[3]}
                # Build translation buckets
                buckets = {"pdt": [], "kjv": [], "sv1888": [], "greeksblgnt": [], "hebrew_oshb": []}
                for r in rows:
                    buckets.setdefault(r[3] or "other", []).append({
                        "word_id": r[0],
                        "surface": r[1],
                        "strongs_id": r[2],
                    })
                # PDT edges to Greek/Hebrew via Strong's
                pdt_edges = []
                for w in buckets.get("pdt", []):
                    sid = w.get("strongs_id")
                    if sid and sid in strongs_map:
                        lang = strongs_map[sid].get("language", "")
                        if lang == "greek":
                            pdt_edges.append({"pdt": w["surface"], "strongs_id": sid, "target": "Greek", "lemma": strongs_map[sid].get("lemma")})
                        elif lang == "hebrew":
                            pdt_edges.append({"pdt": w["surface"], "strongs_id": sid, "target": "Hebrew", "lemma": strongs_map[sid].get("lemma")})
                self.send_json(200, {
                    "verse_id": verse_id,
                    "strongs_count": len(strongs_ids),
                    "word_counts": {k: len(v) for k, v in buckets.items() if v},
                    "pdt_trilingual_edges": pdt_edges,
                }, con)
                return

            if api_path.startswith("/search"):
                q = params.get("q", [""])[0].strip()
                limit = int(params.get("limit", ["20"])[0])
                rows = cur.execute(
                    """
                    SELECT w.*, v.book, v.chapter, v.verse, v.text
                    FROM word_occurrences w
                    JOIN verses v ON w.verse_id = v.verse_id AND w.translation_id = v.translation_id
                    WHERE w.surface LIKE ?
                    ORDER BY v.verse_id, w.position
                    LIMIT ?
                    """,
                    (f"%{q}%", limit),
                ).fetchall()
                self.send_json(200, {"results": [dict(r) for r in rows]}, con)
                return

            if api_path == "/original":
                verse_id = (
                    params.get("verse_id", [""])[0] if "verse_id" in params else ""
                )
                translation_id = (
                    params.get("translation_id", [""])[0]
                    if "translation_id" in params
                    else ""
                )
                if not verse_id:
                    self.send_json(400, {"error": "Missing verse_id"}, con)
                    return
                parts = [c for c in verse_id.split("#") if c]
                if len(parts) >= 3:
                    verse_id = "##".join(parts[:3])
                chapter = parts[1] if len(parts) >= 2 else None
                book_prefix = parts[0] if parts else ""
                book_display = None
                if book_prefix:
                    row = cur.execute(
                        "SELECT book FROM book_code_lookup WHERE code=?",
                        (book_prefix,),
                    ).fetchone()
                    book_display = row[0] if row else None
                candidates = []
                if len(parts) >= 3:
                    candidates.append(
                        (f"{book_prefix}##{parts[1]}##{parts[2]}", "greeksblgnt")
                    )
                    candidates.append((f"{book_prefix}##{parts[1]}##{parts[2]}", ""))
                if book_display and chapter:
                    candidates.append(
                        (
                            f"{book_display.upper()}##{chapter}##{parts[2]}",
                            "greeksblgnt",
                        )
                    )
                    candidates.append(
                        (f"{book_display.upper()}##{chapter}##{parts[2]}", "")
                    )
                rows = []
                used_verse_id = verse_id
                used_translation_id = translation_id or "original"
                for cand_verse_id, cand_translation_id in candidates:
                    q = "SELECT position, surface, strongs_id, morphology FROM word_occurrences WHERE verse_id=? ORDER BY position"
                    args = [cand_verse_id]
                    if cand_translation_id:
                        q += " AND translation_id=?"
                        args.append(cand_translation_id)
                    rows = cur.execute(q, args).fetchall()
                    if rows:
                        used_verse_id = cand_verse_id
                        used_translation_id = cand_translation_id or "original"
                        break
                SEP = " "
                surface = "".join(
                    (
                        r[1]
                        + (
                            SEP
                            if r[3]
                            and any(r[3].startswith(p) for p in ("V-", "N-", "A-", "D-", "R-", "C-", "P-", "X-"))
                            else ""
                        )
                    )
                    for r in rows
                ).strip()
                cleaned_rows = []
                for r in rows:
                    morph = r[3]
                    if morph and not morph.startswith('--------') and not morph.startswith('----') and len(morph) < 30:
                        cleaned_rows.append(r)
                    else:
                        cleaned_rows.append((r[0], r[1], r[2], '', r[4] if len(r) > 4 else None))
                
                result = {
                    "verse_id": used_verse_id,
                    "translation_id": used_translation_id,
                    "surface": surface,
                    "words": [
                        {
                            "position": r[0],
                            "surface": r[1],
                            "strongs_id": r[2],
                            "morphology": r[3] or '',
                        }
                        for r in cleaned_rows
                    ],
                    "language": "hebrew" if any(r[3] and r[3].startswith('H') for r in rows) else "greek",
                    "direction": "rtl" if any(r[3] and r[3].startswith('H') for r in rows) else "ltr",
                }
                self.send_json(200, result, con)
                return

            if api_path.startswith("/strongs"):
                strongs_id = ""
                if api_path.startswith("/strongs/"):
                    strongs_id = api_path[len("/strongs/"):].strip()
                elif "id" in params:
                    strongs_id = params.get("id", [""])[0].strip()
                if not strongs_id:
                    self.send_json(400, {"error": "Missing Strong's ID"}, con)
                    return
                strongs_id = strongs_id.upper()
                lexicon_row = cur.execute(
                    "SELECT id, language, lemma, definition, root_form, cross_refs FROM lexicon_entries WHERE id=?",
                    (strongs_id,),
                ).fetchone()
                if not lexicon_row:
                    self.send_json(404, {"error": "Not found", "strongs_id": strongs_id}, con)
                    return
                usage_rows = cur.execute(
                    """
                    SELECT w.word_id, w.verse_id, w.translation_id, w.surface, w.position,
                           v.book, v.chapter, v.verse, v.text
                    FROM word_occurrences w
                    JOIN verses v ON v.verse_id = w.verse_id AND v.translation_id = w.translation_id
                    WHERE w.strongs_id=?
                    ORDER BY v.verse_id, w.position
                    LIMIT 20
                    """,
                    (strongs_id,),
                ).fetchall()
                occurrences = []
                for r in usage_rows:
                    occurrences.append({
                        "word_id": r[0],
                        "verse_id": r[1],
                        "translation": r[2],
                        "position": r[4],
                        "reference": f"{r[5]} {r[6]}:{r[7]}",
                        "text": r[8],
                    })
                result = {
                    "strongs_id": lexicon_row[0],
                    "language": lexicon_row[1],
                    "lemma": lexicon_row[2],
                    "definition": lexicon_row[3],
                    "root": lexicon_row[4],
                    "cross_refs": lexicon_row[5],
                    "occurrences": len(usage_rows),
                    "verses": occurrences,
                }
                self.send_json(200, result, con)
                return

            if api_path == "/wordcards":
                verse_id = params.get("verse_id", [""])[0].strip()
                if not verse_id:
                    self.send_json(400, {"error": "Missing verse_id"}, con)
                    return
                parts = [c for c in verse_id.split("#") if c]
                if len(parts) >= 3:
                    verse_id = "##".join(parts[:3])
                rows = cur.execute(
                    """
                    SELECT w.position, w.surface, w.strongs_id, w.morphology, l.lemma, l.definition, l.language
                    FROM word_occurrences w
                    LEFT JOIN lexicon_entries l ON l.id = w.strongs_id
                    WHERE w.verse_id=?
                    ORDER BY w.position
                    """,
                    (verse_id,),
                ).fetchall()
                cards = []
                for r in rows:
                    cards.append(
                        {
                            "position": r[0],
                            "word": r[1],
                            "strongs_id": r[2],
                            "morphology": r[3],
                            "lemma": r[4],
                            "definition": r[5],
                            "language": r[6],
                        }
                    )
                self.send_json(200, {"verse_id": verse_id, "cards": cards}, con)
                return

            self.send_json(404, {"error": "Not found", "path": path}, con)
        except Exception as e:
            self.send_json(500, {"error": str(e)}, con)
        finally:
            if con:
                try:
                    con.close()
                except Exception:
                    pass

    def _export_excel(self, con, cur):
        query = self.path.split("?", 1)[1] if "?" in self.path else ""
        params = urllib.parse.parse_qs(query)
        verse_id = params.get("verse_id", [""])[0]
        book_prefix = params.get("book", [""])[0]
        chapter = params.get("chapter", [""])[0]
        whole_bible = str(params.get("whole_bible", ["0"])[0]).lower() in (
            "1",
            "true",
            "yes",
        )
        word_q = None
        word_args = None
        align_q = None
        align_args = None
        scope_filter = "verse"

        if verse_id:
            parts = [c for c in verse_id.split("#") if c]
            if len(parts) >= 3:
                verse_id = "##".join(parts[:3])
            word_q = "SELECT verse_id, position, surface, strongs_id, morphology, translation_id FROM word_occurrences WHERE verse_id=? ORDER BY position"
            word_args = [verse_id]
            align_q = "SELECT id, source_word_id, target_word_id, source_translation, target_translation, strongs_id, alignment_type, quality_score, pos_compatible, alignment_notes, verse_id, sentence_id FROM alignment_edges WHERE verse_id=? ORDER BY id"
            align_args = [verse_id]
            scope_filter = "verse"
        elif whole_bible:
            word_q = "SELECT verse_id, position, surface, strongs_id, morphology, translation_id FROM word_occurrences ORDER BY verse_id, position"
            word_args = []
            align_q = "SELECT id, source_word_id, target_word_id, source_translation, target_translation, strongs_id, alignment_type, quality_score, pos_compatible, alignment_notes, verse_id, sentence_id FROM alignment_edges ORDER BY verse_id, id"
            align_args = []
            scope_filter = "whole_bible"
        elif book_prefix and chapter:
            word_q = "SELECT verse_id, position, surface, strongs_id, morphology, translation_id FROM word_occurrences WHERE verse_id LIKE ? ORDER BY verse_id, position"
            word_args = [f"{book_prefix}##{chapter}##%"]
            align_q = "SELECT id, source_word_id, target_word_id, source_translation, target_translation, strongs_id, alignment_type, quality_score, pos_compatible, alignment_notes, verse_id, sentence_id FROM alignment_edges WHERE verse_id LIKE ? ORDER BY verse_id, id"
            align_args = [f"{book_prefix}##{chapter}##%"]
            scope_filter = "chapter"
        else:
            self.send_json(
                400,
                {"error": "Missing verse_id, book+chapter, or whole_bible=1"},
                con,
            )
            return

        word_rows = cur.execute(word_q, word_args).fetchall()
        align_rows = cur.execute(align_q, align_args).fetchall()
        strongs_ids = sorted({r[3] for r in word_rows if r[3]})
        lex_rows = []
        if strongs_ids:
            placeholders = ",".join("?" for _ in strongs_ids)
            lex_rows = cur.execute(
                f"SELECT id, language, lemma, definition, root_form, cross_refs FROM lexicon_entries WHERE id IN ({placeholders}) ORDER BY id",
                strongs_ids,
            ).fetchall()
        wb = Workbook()
        ws = wb.active
        ws.title = "word_tokens"
        ws.append(
            [
                "verse_id",
                "position",
                "surface",
                "strongs_id",
                "morphology",
                "translation_id",
            ]
        )
        for r in word_rows:
            ws.append(list(r))
        ws2 = wb.create_sheet("alignments")
        ws2.append(
            [
                "id",
                "source_word_id",
                "target_word_id",
                "source_translation",
                "target_translation",
                "strongs_id",
                "alignment_type",
                "quality_score",
                "pos_compatible",
                "alignment_notes",
                "verse_id",
                "sentence_id",
            ]
        )
        for r in align_rows:
            ws2.append(list(r))
        ws3 = wb.create_sheet("lexicon_entries")
        ws3.append(
            [
                "strongs_id",
                "language",
                "lemma",
                "definition",
                "root_form",
                "cross_refs",
            ]
        )
        for r in lex_rows:
            ws3.append(list(r))
        ws4 = wb.create_sheet("summary")
        ws4.append(["metric", "value"])
        ws4.append(["scope", scope_filter])
        ws4.append(["exported_words", len(word_rows)])
        ws4.append(["exported_alignments", len(align_rows)])
        ws4.append(["unique_strongs", len(strongs_ids)])
        unique_verses = sorted({r[0] for r in word_rows if r[0]})
        ws4.append(["verse_count", len(unique_verses)])

        if whole_bible:
            word_rows = None
            align_rows = None
            strongs_ids = None
            lex_rows = None
            ws5 = wb.create_sheet("verses")
            ws5.append(
                ["book_order", "book", "chapter", "verse", "verse_id", "text"]
            )
            book_order = {
                name: idx for idx, name in enumerate(CANONICAL_BOOKS, start=1)
            }
            rows5 = cur.execute(
                "SELECT book, chapter, verse, verse_id, text FROM verses ORDER BY book, chapter, verse"
            ).fetchall()
            ordered = sorted(rows5, key=lambda r: book_order.get(r[0], 9999))
            for idx, r in enumerate(ordered, start=1):
                ws5.append([idx, r[0], r[1], r[2], r[3], r[4]])

            ws6 = wb.create_sheet("book_stats")
            ws6.append(
                [
                    "book_order",
                    "book",
                    "chapter_count",
                    "verse_count",
                    "word_count",
                    "unique_strongs_count",
                ]
            )
            stats = cur.execute(
                """
                SELECT v.book,
                       COUNT(DISTINCT v.chapter) as chapter_count,
                       COUNT(*) as verse_count
                FROM verses v
                GROUP BY v.book
                ORDER BY v.book
                """
            ).fetchall()
            for idx, r in enumerate(stats, start=1):
                book_name = r[0]
                place = book_order.get(book_name, idx)
                ws6.append([place, book_name, r[1], r[2], "", ""])

            ws7 = wb.create_sheet("lexicon_entries")
            ws7.append(
                [
                    "strongs_id",
                    "language",
                    "lemma",
                    "definition",
                    "root_form",
                    "cross_refs",
                ]
            )
            lex_rows = cur.execute(
                "SELECT id, language, lemma, definition, root_form, cross_refs FROM lexicon_entries ORDER BY id"
            ).fetchall()
            for r in lex_rows:
                ws7.append(list(r))

        import io

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        payload = buf.read()
        name = "bible_export.xlsx" if whole_bible else "word_cards_export.xlsx"
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Disposition", f'attachment; filename="{name}"')
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def send_json(self, status_code, data, con=None):
        payload = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def log_message(self, format, *args):
        pass


if __name__ == "__main__":
    server = HTTPServer(("0.0.0.0", PORT), BibleAPIHandler)
    print(f"Bible Mapping Viewer+API running on http://0.0.0.0:{PORT}")
    print(f"Database: {DB_PATH}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.shutdown()
